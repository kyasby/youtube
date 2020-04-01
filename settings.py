# -*- coding: utf-8 -*-
import sys
import time
import pandas as pd
import datetime
from main import make_ch_date
from main import fetch_vd_data
from main import activate_log
import logging
from logging import getLogger, StreamHandler, Formatter,FileHandler
import openpyxl
import tkinter as tk
from tkinter import messagebox

logger = getLogger("logger").getChild("sub")

def setting():

    #settingファイルの設定
    st = pd.read_excel('settings.xlsx')
    CH_NAME           = st.iloc[1]
    CH_ID             = st.iloc[2]
    WHETHER_GET       = st.iloc[3]
    START_DT          = st.iloc[4]
    LAST_DT           = st.iloc[5]
    SIZE_FETCHED_VD   = st.iloc[6]
    API_KEY           = st.iloc[8]
    REMAIN_API        = st.iloc[9]
    UPDATE_AT         = st.iloc[10]
    ROW_NUM           = st.iloc[12]

    #エクセルを読み込み
    book = openpyxl.load_workbook('settings.xlsx')
    sheet = book['シート1']
    
    #日付
    today = str(datetime.date.today())#L77でも使う
    pst_time = datetime.datetime.today() + datetime.timedelta(hours=-17)

    activate_log()
    consumed_api = 0
    
    logger.info("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~開始~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    
    #列ごとに繰り返し
    for row in range(1,ROW_NUM[1]+1):
        if WHETHER_GET[row].lower() ==  "y":
            #入力ミス探知
            # if len(START_DT[row]) != 10:
            #     tk.Tk().withdraw()
            #     logger.info("~~~~~~~~~~~~~~~~~~~~~~~~エクセルの入力ミスで終了(日付)~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            #     messagebox.showwarning("エラー", CH_NAME[row]+"の日付を1998-10-01/nowのように入力してください。")
            #     sys.exit(0)
            if len(LAST_DT[row]) != 10 and len(LAST_DT[row]) != 3:
                tk.Tk().withdraw()
                logger.info("~~~~~~~~~~~~~~~~~~~~~~~~エクセルの入力ミスで終了(日付)~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
                messagebox.showwarning("エラー", CH_NAME[row]+"の日付を1998-10-01/nowのように入力してください。")
                sys.exit(0)
            if len(CH_ID[row]) != 24:
                tk.Tk().withdraw()
                logger.info("~~~~~~~~~~~~~~~~~~~~~~~~エクセルの入力ミスで終了(チャンネルID)~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
                messagebox.showwarning("エラー", CH_NAME[row]+"のチャンネルIDを確認してください。")
                sys.exit(0)

    for row in range(1,ROW_NUM[1]+1):
        dates = []
        
        if WHETHER_GET[row].lower() ==  "y":
            ch_name = CH_NAME[row]
            logger.info("~~~~~~~~~~~~~~~~~~~~チャンネル：" + ch_name + "を開始~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            CHANNEL_ID = CH_ID[row]
            api_key = API_KEY[row]
            ramain_api = REMAIN_API[row]
            
            try:#最新のXX件を取得
                NUM = int(SIZE_FETCHED_VD[row])
                # LAST_DT[row] = "now"
                # START_DT[row] = "1998-10-01"
            except:
                NUM = 99999999

            #start_dtの設定
            if START_DT[row].lower() == "30days":
                start_dt = str(datetime.date.today()+ datetime.timedelta(days=-30)) + "T00:00:00.000Z"
            else:
                start_dt     = START_DT[row] + "T00:00:00.000Z"
            

            #middleの設定
            try:
            #前データの最新の日付に１秒足した時間を取得
                logger.info("前回の最新の日付を取得")
                latest_dt = pd.read_csv(ch_name + '\\' + CH_ID[row]+'.csv').iloc[0][3]
                #1998.10.01T10:10:01Z => 10:10:01
                tmp  = str(datetime.datetime.fromisoformat(latest_dt.replace("Z","")))[11:]
                #10:10:01=>10:10:02
                M  = str((datetime.datetime.strptime(tmp, '%H:%M:%S') + datetime.timedelta(seconds=1)).time())
                #1998.10.01T10:10:02.000Z
                middle = latest_dt[:10]+"T"+M+".000Z"
            except:
                logger.info("初回のようだ。")
                if START_DT[row].lower() == "30days":
                    middle = str(datetime.date.today() + datetime.timedelta(days=-30)) + "T00:00:00.000Z"
                else:
                    middle = START_DT[row] + "T00:00:00.000Z"

            #last_dtの設定
            if LAST_DT[row].lower() == "now":
                last_dt    = str(datetime.date.today() + datetime.timedelta(days=1)) + "T00:00:00.000Z"
            else:
                last_dt    = LAST_DT[row] + "T10:00:00.000Z"
            logger.info("日付定義完了:" + middle + "から" + last_dt + "のデータを取得します。")

            middle_check  = datetime.datetime.fromisoformat(middle.replace('Z', ''))
            last_dt_check = datetime.datetime.fromisoformat(last_dt.replace('Z', ''))
            # 2020-03-23T00:00:00.000Z last
            # 2020-03-23T00:00:00.000 replace
            # 2020-03-23 00:00:00 check

            flag = False
            if middle_check <= last_dt_check:#last_dtが前データよりも最新の日付を指定しているならば。
                ch_result = make_ch_date(CHANNEL_ID, middle, last_dt, ch_name, api_key)############################新動画を取得
                flag = ch_result[0]
                consumed_api += ch_result[1]

            if flag == True:
                logger.error('エラーなので動画の取得を中止します。')
            else:
                logger.info('読み込みCSVを最新に更新')
                ch = pd.read_csv(ch_name + '\\' + CHANNEL_ID+'.csv')
                start_dt = datetime.datetime.fromisoformat(start_dt.replace('Z', ''))
                last_dt = datetime.datetime.fromisoformat(last_dt.replace('Z', '')) + datetime.timedelta(days=1)
                logger.info(str(start_dt) +'から'+str(last_dt)+'までのビデオの取得開始')
                consumed_api += fetch_vd_data(ch, start_dt, last_dt, ch_name, NUM, api_key) #ch = pd.read_csv(チャンネルID.csv')
                logger.info("~~~~~~~~~~~~~~~~~~~~チャンネル：" + ch_name + "を終了~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    
    #if str(UPDATE_AT[1]).replace(" 00:00:00","") != str(pst_time.date()):#太平洋時間が変わっている場合
    #    sheet.cell(row=11, column=(2)).value=10000 - consumed_api
    #    sheet.cell(row=12, column=(2)).value= pst_time.date()
    #    sheet.cell(row=13, column=(2)).value= consumed_api
    #else:#太平洋時間が変わっていない場合
    #    sheet.cell(row=11, column=(2)).value = sheet.cell(row=11, column=(2)).value - consumed_api
    #    sheet.cell(row=13, column=(2)).value = consumed_api
    #book.save('settings.xlsx')

    logger.info("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~終了~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    logger.info("消費API")
    logger.info(consumed_api)
    
    


if __name__ == "__main__":
    try:
        setting()
    except:
        logger.error("hmm", exc_info=True)